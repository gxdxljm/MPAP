# -*- coding: utf-8 -*-
"""
Adapted from ds4dm/learn2branch (https://github.com/ds4dm/learn2comparenodes).
Modified for MPAP (Multi-layer Perceptron with Attention Pooling) node selection framework under the same MIT License.
"""
import os
import re
import numpy as np
from openpyxl import Workbook, load_workbook
from scipy.stats import gmean
from scipy.stats import gstd
import pyscipopt.scip as sp
from node_selection.recorders import CompFeaturizerSVM, CompFeaturizer, LPFeatureRecorder
# from node_selection.node_selectors import (CustomNodeSelector,
#                                            OracleNodeSelectorAbdel,
#                                            OracleNodeSelectorEstimator_SVM,
#                                            OracleNodeSelectorEstimator_RankNet,
#                                            OracleNodeSelectorEstimator)      # 源码内容: 佳明注释,我们不再需要SVM和GNN
from node_selection.node_selectors import (CustomNodeSelector,
                                           OracleNodeSelectorAbdel,
                                           OracleNodeSelectorEstimator_RankNet,
                                           )
# from learning.utils import normalize_graph

def distribute(n_instance, n_cpu):
    if n_cpu == 1:
        return [(0, n_instance)]
    
    k = n_instance //( n_cpu -1 )
    r = n_instance % (n_cpu - 1 )
    res = []
    for i in range(n_cpu -1):
        res.append( ((k*i), (k*(i+1))) )
    
    res.append(((n_cpu - 1) *k ,(n_cpu - 1) *k + r ))
    return res


def get_nodesels2models(nodesels, instance, problem, normalize, device):
    
    res = dict()
    nodesels2nodeselectors = dict()
    
    for nodesel in nodesels:
        
        model = sp.Model()

        model.hideOutput()  # 禁止日志显示

        model.readProblem(instance)
        model.setIntParam('randomization/permutationseed', 5)  # 固定变量/约束重排顺序, 确保实验可复现，消除随机噪声干扰
        model.setIntParam('randomization/randomseedshift',5)   # 固定全局随机种子, 确保实验可复现，消除随机噪声干扰
        model.setParam('constraints/linear/upgrade/logicor',0)
        model.setParam('constraints/linear/upgrade/indicator',0)
        model.setParam('constraints/linear/upgrade/knapsack', 0)
        model.setParam('constraints/linear/upgrade/setppc', 0)
        model.setParam('constraints/linear/upgrade/xor', 0)
        model.setParam('constraints/linear/upgrade/varbound', 0)

        # 佳明新增,避免单个算例过长时间的测试
        model.setParam('limits/time',700.0)  # 严格控制测试时间
        
        comp = None
        
        # if not re.match('default*', nodesel):  # 源码内容
        if nodesel != "default":                 # 佳明修改,确保原意不变的同时,简化理解
            try:
                comp_policy, sel_policy, other = nodesel.split("_")  # 例如: "gnn_dummy_nprimal=2" -> ["gnn", "dummy", "nprimal=2"]
            # except:           # 源码内容
            except ValueError:  # 佳明修改,确保原意不变的同时,简化理解
                comp_policy, sel_policy = nodesel.split("_")
                


            # if comp_policy == 'gnn':
            #     comp_featurizer = CompFeaturizer()
            #
            #     feature_normalizor = normalize_graph if normalize else lambda x: x  # 是否归一化,normalize=True执行图归一化
            #
            #     n_primal = int(other.split('=')[-1])
            #
            #
            #     comp = OracleNodeSelectorEstimator(problem,
            #                                        comp_featurizer,
            #                                        device,
            #                                        feature_normalizor,
            #                                        use_trained_gnn=True,
            #                                        sel_policy=sel_policy,
            #                                        n_primal=n_primal)
            #     fr = LPFeatureRecorder(model, device)
            #     comp.set_LP_feature_recorder(fr)

            ##  源码内容,佳明注释: 我们不需要支持向量机参与比较
            # elif comp_policy == 'svm':
            #     comp_featurizer = CompFeaturizerSVM(model)
            #     n_primal = int(other.split('=')[-1])
            #     comp = OracleNodeSelectorEstimator_SVM(problem, comp_featurizer, sel_policy=sel_policy, n_primal=n_primal)
                
            if comp_policy == 'ranknet':  # 多层感知机MLP,需要用到getHeHeaumeEisnerFeatures函数提取He等人的特征
                comp_featurizer = CompFeaturizerSVM(model)
                dir_path = os.path.dirname(instance)
                pinfo_path = os.path.join(dir_path, f'{problem}_Pinfo.csv')
                comp_featurizer._load_var_info(pinfo_path)                   # 在特征提取器注册物理信息的相关变量
                n_primal = int(other.split('=')[-1])  # n_primal表示神经网络启动推理的位置, 例如: n_primal=2表示从第2次primal_bound更新以后,再开始神经网络推理
                comp = OracleNodeSelectorEstimator_RankNet(problem, comp_featurizer, device, sel_policy=sel_policy, n_primal=n_primal)


            elif comp_policy == 'expert':  # oracle专家
                comp = OracleNodeSelectorAbdel('optimal_plunger', optsol=0, inv_proba=0)
                optsol = model.readSolFile(instance.replace(".lp", ".sol"))  # 读取最优解文件,得到pyscipopt.Solution对象,,用于生成oracle路径
                comp.setOptsol(optsol)

                ## 佳明注释: 辅助理解optsol的含义
                # print(optsol.getObjVal())                        # 目标值
                # var_dict = {v.name: v for v in model.getVars()}  # 最优解
                # relevant_vars = ['uit(1,2)']
                # for var_name in relevant_vars:
                #     var = var_dict[var_name]
                #     val = optsol.getVal(var)
                #     print(f"var: {var}, value: {val:.2f}")

            else:  # 佳明注释: 指定的某一种节点选择策略,全程使用文献作者自定义该策略的细节, 暂时认为也无需实现
                comp = CustomNodeSelector(comp_policy=comp_policy, sel_policy=sel_policy)

            model.includeNodesel(comp, nodesel, 'testing',  536870911,  536870911)
        
        else:  # SCIP默认的节点选择策略,人为设定某种节点选择规则的优先级
            _, nsel_name, priority = nodesel.split("_")
            # assert(nsel_name in ['estimate', 'dfs', 'bfs']) #to do add other default methods                    # 源码内容,针对SCIP7.X版本
            assert (nsel_name in ['estimate', 'bfs', 'hybridestim', 'restartdfs', 'uct', 'dfs', 'breadthfirst'])  # 佳明修改 针对SCIP10.0.0版本
            priority = int(priority)
            model.setNodeselPriority(nsel_name, priority)  # pyscipopt接口函数: 设定节点选择策略优先级
            

            
        
        res[nodesel] = model
        nodesels2nodeselectors[nodesel] = comp
        
        
        
            
    return res, nodesels2nodeselectors


def get_record_file(problem, nodesel, instance, save_dir):
    # save_dir = os.path.join(os.path.abspath(''),  f'stats/{problem}/{nodesel}/')  # 源码内容
    save_dir_tmp = os.path.join(save_dir, f'{problem}/{nodesel}/')
    
    # try:                         # 源码内容
    #     os.makedirs(save_dir)
    # except FileExistsError:
    #     ""
    try:
        os.makedirs(save_dir_tmp)  # 不允许重复创建文件夹,必须手动重命名旧文件夹再重新运行程序
    except FileExistsError:
        ""
        
    # instance = str(instance).split('/')[-1]   # 源码内容
    instance = os.path.basename(str(instance))  # 佳明修改,适配windows系统,从完整的文件路径中提取文件名称
    file = os.path.join(save_dir_tmp, instance.replace('.lp', '.csv'))  # 将后缀名由.lp改为.csv,再拼贴基础路径得到结果文件的完整路径
    return file


def record_stats_instance(problem, nodesel, model, instance, nodesel_obj, save_dir, verbose=False):
    nnode = model.getNNodes()
    nlps = model.getNLPs()
    nlpit = model.getNLPIterations()
    current_seed = model.getParam('randomization/permutationseed')
    stime =round(model.getSolvingTime(),1)
    gap = round(model.getGap(), 6)
    status = model.getStatus()
    final_primal = round(model.getObjVal(),1)
    final_dual = round(model.getDualbound(),1)

    # 2026.3.5 佳明注释
    # nodesel是用户外部输入的待测试的节点策略的名称
    # nodesel_obj是节点选择器,定义了具体的comp和sel方法
    if nodesel_obj != None:    
        comp_counter = nodesel_obj.comp_counter
        sel_counter = nodesel_obj.sel_counter
    else:
        comp_counter = sel_counter = -1
    
    if re.match('gnn*', nodesel):
        init1_time = nodesel_obj.init_solver_cpu
        init2_time = nodesel_obj.init_cpu_gpu
        fe_time = nodesel_obj.fe_time
        fn_time = nodesel_obj.fn_time
        inference_time = nodesel_obj.inference_time
        inf_counter = nodesel_obj.inf_counter
        
    else:
        init1_time, init2_time, fe_time, fn_time, inference_time, inf_counter = -1, -1, -1, -1, -1, -1
    
    if re.match('ranknet*', nodesel):
        inf_counter = nodesel_obj.inf_counter
        fe_time = nodesel_obj.fea_time
        inference_time = nodesel_obj.inf_time
        comp_counter = nodesel_obj.comp_counter
        sel_counter = nodesel_obj.sel_counter

    if re.match('svm*', nodesel) or re.match('expert*', nodesel):
        comp_counter = nodesel_obj.comp_counter
        sel_counter = nodesel_obj.sel_counter

    ##  节点数 总求解时间 节点比较总次数 节点选择总次数 初始化用时1 初始化用时2 二部图提取时间 二部图归一化时间 总推理时间 推理总次数
    ## 源码内容, 以不包含表头的方式存储计算结果, 由于不够直观, 被本人弃用
    # file = get_record_file(problem, nodesel, instance, save_dir)
    # np.savetxt(file, np.array([nnode, stime, comp_counter, sel_counter, init1_time, init2_time, fe_time, fn_time, inference_time, inf_counter]), delimiter=',')

    # 2026.3.4 佳明新增 由于Abdel的源码没有用1个文件汇总保存所有.lp文件的求解结果,因此本人补充该功能
    result_row = [
        nodesel,        # str
        instance,       # str (e.g., "case118_67")
        current_seed,   # int
        final_primal,  # float
        final_dual,    # float
        nnode,  # int
        nlps,  # int
        nlpit,  # int
        stime,  # float
        gap,  # float
        status,  # str (e.g., "optimal", "timelimit")
        comp_counter,  # int
        sel_counter,  # int
        inf_counter,  # int
        fe_time ,     # float
        inference_time, # float
    ]
    # --- 构造 Excel 文件路径 ---
    excel_filename = f"{problem}-test-results.xlsx"
    excel_path = os.path.join(save_dir, excel_filename)
    # --- 表头（仅在新建时写入）---
    headers = [
        'node_sel', 'instance', 'current_seed', 'final_primal', 'final_dual', 'nnodes', 'nlps', 'nlpit',
        'stime', 'gap', 'status', 'comp_counter', 'sel_counter', 'inf_counter',
        'feature_time', 'inf_time'
    ]
    # --- 写入逻辑 ---
    if os.path.exists(excel_path):
        # 文件存在：加载并追加到第一个空行
        workbook = load_workbook(excel_path)
        sheet = workbook.active
    else:
        # 文件不存在：新建并写入表头 + 第一行
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(headers)  # 写表头
    # 追加新数据行
    sheet.append(result_row)
    # 保存并关闭
    workbook.save(excel_path)
    workbook.close()
    if verbose:
        instance_name = os.path.basename(str(instance))
        print(f"Instance: {instance_name}, nodesel: {nodesel}, nnodes: {nnode}, comp: {comp_counter}, time: {stime:.2f}")


def print_infos(problem, nodesel, instance):
    print("------------------------------------------")
    print(f"   |----Solving:  {problem}")
    print(f"   |----Instance: {instance}")
    print(f"   |----Nodesel: {nodesel}")


def solve_and_record_default(problem, instance, verbose, save_dir):
    default_model = sp.Model()

    default_model.hideOutput()  # 禁止日志显示

    default_model.setIntParam('randomization/permutationseed',5)  # 固定变量/约束重排顺序, 确保实验可复现，消除随机噪声干扰
    default_model.setIntParam('randomization/randomseedshift',5)  # 固定全局随机种子, 确保实验可复现，消除随机噪声干扰
    default_model.readProblem(instance)

    # 佳明新增,避免单个算例过长时间的测试
    default_model.setParam('limits/time', 700.0)  # 严格控制测试时间

    if verbose:
        print_infos(problem, 'default', instance)
    
    default_model.optimize()        
    record_stats_instance(problem, 'default', default_model, instance, None, save_dir)


#take a list of nodeselectors to evaluate, a list of instance to test on, and the 
#problem type for printing purposes
def record_stats(nodesels, instances, problem, device, normalize, save_dir, verbose=False, default=True):
    

    for instance in instances:       
        instance = str(instance)
        
        if default and not os.path.isfile(get_record_file(problem,'default', instance, save_dir)):  # 检查是否有旧的运行结果文件
            solve_and_record_default(problem, instance, verbose, save_dir)  # 2026.3.2 佳明注释: 记录SCIP在default设置的结果
        
        
        nodesels2models, nodesels2nodeselectors = get_nodesels2models(nodesels, instance, problem, normalize, device)  # 注册各种节点选择器,设置求解器参数,读入.lp文件
        
        for nodesel in nodesels:  
            
            model = nodesels2models[nodesel]
            nodeselector = nodesels2nodeselectors[nodesel]
                
           #test nodesels
            if os.path.isfile(get_record_file(problem, nodesel, instance, save_dir)): #no need to resolve
                continue
        
            
            if verbose:
                print_infos(problem, nodesel, instance)  # 每个算例的日志
                
            model.optimize()
            record_stats_instance(problem, nodesel, model, instance, nodeselector, save_dir)  # 记录求解结束后,单独记录每一个.lp文件的求解信息(总耗时,总节点数...),节点选择器的信息(节点选择次数,左右节点比较次数...)
    

def get_mean(problem, nodesel, instances, stat_type, save_dir):
    res = []
    n = 0
    means = dict()
    stat_idx = ['nnode', 'time', 'ncomp','nsel', 'init1', 'init2', 'fe', 'fn', 'inf','ninf'].index(stat_type)
    for instance in instances:
        try:
            file = get_record_file(problem, nodesel, instance, save_dir)
            res.append(np.genfromtxt(file)[stat_idx])
            n += 1
            means[str(instance)] = np.genfromtxt(file)[stat_idx]
        except:
            ''
    
    if stat_type in ['nnode', 'time'] :

        mu = np.exp(np.mean(np.log(np.array(res) + 1 )))

        std = np.exp(np.sqrt(np.mean(  ( np.log(np.array(res)+1) - np.log(mu) )**2 )))
    else:
        mu, std = np.mean(res), np.std(res)

    return mu,n, means,  std 


def display_stats(problem, nodesels, instances, min_n, max_n, save_dir, default=False):
    
    print("======================================================")
    print(f'Statistics on {problem} for problem size in [{min_n}, {max_n}]') 
    print("======================================================")
    means_nodes = dict()
    for nodesel in (['default'] if default else []) + nodesels:
        
            
        nnode_mean, n, nnode_means, nnode_dev = get_mean(problem, nodesel, instances, 'nnode', save_dir)
        time_mean, _, _, time_dev  =  get_mean(problem, nodesel, instances, 'time', save_dir)
        ncomp_mean = get_mean(problem, nodesel, instances, 'ncomp', save_dir)[0]
        nsel_mean = get_mean(problem, nodesel, instances, 'nsel', save_dir)[0]
        
        
        means_nodes[nodesel] = nnode_means
        
    
        print(f"  {nodesel} ")
        print(f"      Mean over n={n} instances : ")
        print(f"        |- B&B Tree Size   :  {nnode_mean:.2f}  ± {nnode_dev:.2f}")
        if re.match('gnn*', nodesel):
            in1_mean = get_mean(problem, nodesel, instances, 'init1', save_dir)[0]
            in2_mean = get_mean(problem, nodesel, instances, 'init2', save_dir)[0]
            print(f"        |- Presolving A,b,c Feature Extraction Time :  ")
            print(f"           |---   Init. Solver to CPU:           {in1_mean:.2f}")
            print(f"           |---   Init. CPU to GPU   :           {in2_mean:.2f}")
        print(f"        |- Solving Time    :  {time_mean:.2f}  ± {time_dev:.2f}")
        
        #print(f"    Median number of node created : {np.median(nnodes):.2f}")
        #print(f"    Median solving time           : {np.median(times):.2f}""
    
    
                
        if re.match('gnn*', nodesel):
            fe_mean = get_mean(problem, nodesel, instances, 'fe', save_dir)[0]
            fn_mean = get_mean(problem, nodesel, instances, 'fn', save_dir)[0]
            inf_mean = get_mean(problem, nodesel, instances, 'inf', save_dir)[0]
            print(f"           |---   On-GPU Feature Updates:        {fe_mean:.2f}")
            print(f"           |---   Feature Normalization:         {fn_mean:.2f}")
            print(f"           |---   Inference     :                {inf_mean:.2f}")
            
        if not re.match('default*', nodesel):
            print(f"        |- nodecomp calls  :  {ncomp_mean:.0f}")
            if re.match('gnn*', nodesel) or re.match('svm*', nodesel) or re.match('expert*', nodesel) or re.match('ranknet*', nodesel):
                inf_counter_mean = get_mean(problem, nodesel, instances, 'ninf', save_dir)[0]
                print(f"           |---   inference nodecomp calls:      {inf_counter_mean:.0f}")
            print(f"        |- nodesel calls   :  {nsel_mean:.0f}")
        print("-------------------------------------------------")
        
    return means_nodes
     
     
    
